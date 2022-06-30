from pathlib import Path
from datetime import date
import time

today = date.today()
month = today.month
year = today.year

##used for live running
path_to_file = "/mnt/uCloud/Data/PPG/Reportes/Automatico/{}/{}/ReporteAutomatico-{}.xlsx".format(year,month,today)
#next line just for test purposes. please change when is in production
#path_to_file = "C:/Users/jfranco/OneDrive /Rivas/AutoReport/uCloud/Data/PPG/Reportes/Automatico/{}/{}/ReporteAutomatico-{}.xlsx".format(year,month,today)
path = Path(path_to_file)

for i in range(10):
    time.sleep(5)
    while True:
        if path.is_file():
            print(f'deja vuelvo a checar si hay cambios')
            break
        else:
            print(f'ahi te va un nuevo reporte!')
            # importar las librerias
            import os
            import glob
            from datetime import timedelta
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.comments import Comment
            import warnings

            warnings.simplefilter("ignore")

            # saca el nombre del ultimo archivo subido
            list_of_files = glob.glob(
                '/mnt/uCloud/Data/PPG/Reportes/Automatico/ReportePOs_IRs/*.xlsx')  # * means all if need specific format then *.csv
                #'C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/PPG/Reportes/Automatico/ReportePOs_IRs/*.xlsx')  # * means all if need specific format then *.csv #DEV TEST
            latest_file = max(list_of_files, key=os.path.getctime)


            ## ubicaciones
            os.chdir("/mnt/uCloud/Data/")
            ##os.chdir("C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/")
            path = "/mnt/uCloud/Data/Databases/AutoReportDaily/PPG_BD_automatico.xlsx"
            ##path = "C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/Databases/AutoReportDaily/PPG_BD_automatico.xlsx"
            #export = "/mnt/uCloud/Data/PPG/Reportes/Automatico/ReporteAutomatico-{}.xlsx".format(today) ### este ya no lo necesito
            ##c_ClaveProdServCP = "C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/PPG/Reportes/Automatico/Catalogos/c_ClaveProdServCP.xlsx"
            c_ClaveProdServCP = "/mnt/uCloud/Data/PPG/Reportes/Automatico/Catalogos/c_ClaveProdServCP.xlsx"
            ##c_MaterialPeligroso = "C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/PPG/Reportes/Automatico/Catalogos/c_MaterialPeligroso.xlsx"
            c_MaterialPeligroso = "/mnt/uCloud/Data/PPG/Reportes/Automatico/Catalogos/c_MaterialPeligroso.xlsx"
            ##UN_Importaciones = "C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/PPG/Reportes/Automatico/Catalogos/UN_Importaciones.xlsx"
            UN_Importaciones = "/mnt/uCloud/Data/PPG/Reportes/Automatico/Catalogos/UN_Importaciones.xlsx"
            ##Catalogo = "C:/Users/jfranco/OneDrive - lamodernaUSA/Rivas/AutoReport/uCloud/Data/PPG/Reportes/Automatico/Catalogos/Catalogo.xlsx"
            Catalogo = "/mnt/uCloud/Data/PPG/Reportes/Automatico/Catalogos/Catalogo.xlsx"
            db = latest_file
            export = path_to_file

            # Abrir el archivos
            wb_obj = openpyxl.load_workbook(path)
            wb_obj_db = openpyxl.load_workbook(db)
            wb_obj_db3 = openpyxl.load_workbook(c_ClaveProdServCP)
            wb_obj_db4 = openpyxl.load_workbook(c_MaterialPeligroso)
            wb_obj_db5 = openpyxl.load_workbook(UN_Importaciones)
            wb_obj_db6 = openpyxl.load_workbook(Catalogo)

            # seleccionar hoja activa
            sheet_obj = wb_obj.active
            max_col = sheet_obj.max_column
            max_r = sheet_obj.max_row

            sheet_obj_db = wb_obj_db.active
            max_col_db = sheet_obj_db.max_column
            max_r_db = sheet_obj_db.max_row

            sheet_obj_db3 = wb_obj_db3.active
            max_col_db3 = sheet_obj_db3.max_column
            max_r_db3 = sheet_obj_db3.max_row

            sheet_obj_db4 = wb_obj_db4.active
            max_col_db4 = sheet_obj_db4.max_column
            max_r_db4 = sheet_obj_db4.max_row

            sheet_obj_db5 = wb_obj_db5.active
            max_col_db5 = sheet_obj_db5.max_column
            max_r_db5 = sheet_obj_db5.max_row

            sheet_obj_db6 = wb_obj_db6.active
            max_col_db6 = sheet_obj_db6.max_column
            max_r_db6 = sheet_obj_db6.max_row

            # propiedades de la hoja
            comment = Comment("Desarrollado por jfranco@rivaslozano.com", "Ing. Pepe Franco")
            comment.width = 300
            comment.height = 50
            sheet_obj["A1"].comment = comment

            # Color and format to the header of each column
            for i in range(1, max_col + 19):
                cell_obj = sheet_obj.cell(row=1, column=i)
                cell_obj.font = Font(b=True, size=13, color="00FFFFFF")
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                cell_obj.fill = PatternFill("solid", fgColor="444444")

            ###sheet_obj["A3"].font = big_red_text
            sheet_obj.insert_rows(1)
            # sheet_obj["E2"].fill = PatternFill("solid", fgColor="0000CCFF")
            sheet_obj["P2"].fill = PatternFill("solid", fgColor="00008000")
            sheet_obj["Q2"].fill = PatternFill("solid", fgColor="00008000")
            sheet_obj["R2"].fill = PatternFill("solid", fgColor="00008000")
            sheet_obj["S2"].fill = PatternFill("solid", fgColor="00008000")
            sheet_obj["T2"].fill = PatternFill("solid", fgColor="00008000")
            sheet_obj["A2"] = 'CONTROL ID'
            sheet_obj["B2"] = 'FABRICANTE'
            sheet_obj["C2"] = 'ORDEN DE COMPRA'
            sheet_obj["D2"] = 'PRODUCTO'
            sheet_obj["E2"] = 'CON FACTURA'
            sheet_obj["F2"] = 'CANTIDAD'
            sheet_obj["G2"] = 'TAMANO'
            sheet_obj["H2"] = 'PESO'
            sheet_obj["I2"] = 'TIPO DE PESO'
            sheet_obj["J2"] = 'LOTE'
            sheet_obj["K2"] = 'COA'
            sheet_obj["L2"] = 'TRANSPORTE'
            sheet_obj["M2"] = 'CAJA'
            sheet_obj["N2"] = 'LLEGADA A LAREDO'
            sheet_obj["O2"] = 'DESTINO/RL'
            sheet_obj["P2"] = 'DIAS EN LAREDO'
            sheet_obj["Q2"] = 'FECHA DE ENTREGA'
            sheet_obj["R2"] = 'USUARIO'
            sheet_obj["S2"] = 'DESTINO/PPG-PO'
            sheet_obj["T2"] = 'OBSERVACIONES'
            sheet_obj["U2"] = 'CLASIFICACION-PRODUCTO'
            sheet_obj["V1"] = 'SAT'
            sheet_obj["V2"] = 'DESCRIPCION'
            sheet_obj["W2"] = 'MATERIAL PELIGROSO'
            sheet_obj["X2"] = 'CANTIDAD'
            sheet_obj["Y2"] = 'MATERIAL PELIGROSO'
            sheet_obj["Z2"] = 'DESCRIPCION'
            sheet_obj["AA2"] = 'CLASE-DIV'
            sheet_obj["AB2"] = 'PESO-KGS'
            sheet_obj["AC2"] = 'VALOR-MERCANCIA'
            sheet_obj["AD2"] = 'MONEDA'
            sheet_obj["AE2"] = 'PEDIMENTO'
            sheet_obj["AF2"] = 'FRACCION-ARANCELARIA'
            sheet_obj["AG2"] = 'DESCRIPCION ARANCELARIA'
            sheet_obj["AH2"] = 'TIPO DE EMBALAJE'


            ## Extras
            sheet_obj["E1"] = 'OPERACION DIARIA - RIVAS LOZANO'
            for head1 in range(1, max_col + 5):
                cell_obj = sheet_obj.cell(row=1, column=head1)
                cell_obj.font = Font(b=True, size=24, color="ffffff")
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                cell_obj.fill = PatternFill("solid", fgColor="960000")

            ## Color en SAT HEADER
            for head11 in range(21, 24):
                cell_obj = sheet_obj.cell(row=1, column=head11)
                cell_obj.font = Font(b=True, size=24, color="ffffff")
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                cell_obj.fill = PatternFill("solid", fgColor="4bacc6")
            for head21 in range(21, 24):
                cell_obj = sheet_obj.cell(row=2, column=head21)
                cell_obj.fill = PatternFill("solid", fgColor="4bacc6")
            for head31 in range(32, 34):
                cell_obj = sheet_obj.cell(row=1, column=head31)
                cell_obj.fill = PatternFill("solid", fgColor="4bacc6")
                cell_obj = sheet_obj.cell(row=2, column=head31)
                cell_obj.fill = PatternFill("solid", fgColor="4bacc6")

            # Format of Width
            sheet_obj.column_dimensions['B'].width = 35
            sheet_obj.column_dimensions['C'].width = 25
            sheet_obj.column_dimensions['O'].width = 35
            sheet_obj.column_dimensions['P'].width = 35
            sheet_obj.column_dimensions['Q'].width = 35
            sheet_obj.column_dimensions['R'].width = 35
            sheet_obj.column_dimensions['S'].width = 50
            sheet_obj.column_dimensions['T'].width = 70
            sheet_obj.column_dimensions['U'].width = 35
            sheet_obj.column_dimensions['V'].width = 35
            sheet_obj.column_dimensions['W'].width = 35
            sheet_obj.column_dimensions['X'].width = 35
            sheet_obj.column_dimensions['Y'].width = 35
            sheet_obj.column_dimensions['Z'].width = 35
            sheet_obj.column_dimensions['AA'].width = 35
            sheet_obj.column_dimensions['AB'].width = 35
            sheet_obj.column_dimensions['AC'].width = 35
            sheet_obj.column_dimensions['AD'].width = 35
            sheet_obj.column_dimensions['AE'].width = 35
            sheet_obj.column_dimensions['AF'].width = 30
            sheet_obj.column_dimensions['AG'].width = 70
            sheet_obj.column_dimensions['AH'].width = 35

            # has invoice color
            for v in range(1, max_r + 2):
                c_value = sheet_obj.cell(row=v, column=5).value
                if c_value == "No":
                    cell_obj = sheet_obj.cell(row=v, column=5)
                    cell_obj.font = Font(b=True, size=12, color="00FFFFFF")
                    cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                    cell_obj.fill = PatternFill("solid", fgColor="00003366")
                elif c_value == "Yes":
                    cell_obj = sheet_obj.cell(row=v, column=5, value="Si")
                    cell_obj.font = Font(b=True, size=12)
                    cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                    cell_obj.fill = PatternFill("solid", fgColor="0099CC00")

            ######## FIX COLUMNS#########

            #  formato la columna de Control ID
            for cid in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=cid, column=1)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #  formato la columna de fabricante
            for fab in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=fab, column=2)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #  formato la columna de PO
            for po in range(3, max_r + 2):
                c_value = sheet_obj.cell(row=po, column=3).value
                data1 = str(c_value)
                if data1.isdigit():
                    data1 = int(c_value)
                cell_obj = sheet_obj.cell(row=po, column=3, value=data1)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #  formato la columna de producto
            for prod in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=prod, column=4)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #  formato la columna de tamanos
            for f in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=f, column=7)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #  formato la columna de lote
            for l in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=l, column=10)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            # calculo de fecha de entrega -> fecha de llegada + 2
            for ldo in range(3, max_r + 2):
                c_value = sheet_obj.cell(row=ldo, column=14).value.date()
                end_date = c_value + timedelta(days=2)
                cell_obj = sheet_obj.cell(row=ldo, column=17, value=end_date.strftime("%m/%d/%Y"))

            #  formato la columna de llegada a laredo
            for ll in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=ll, column=14)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            # formato de la columna fecha de entrega
            for ll in range(1, max_r + 2):
                cell_obj = sheet_obj.cell(row=ll, column=17)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #  formato la columna de COA
            for coa in range(1, max_r + 2):
                c_value = sheet_obj.cell(row=coa, column=11).value
                if c_value == "Check":
                    cell_obj = sheet_obj.cell(row=coa, column=11, value="Incluido")
                    cell_obj.font = Font(b=True, size=12)
                    cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                    cell_obj.fill = PatternFill("solid", fgColor="0099CC00")

            # dias en laredo
            for v in range(2, max_r + 2):
                cell_obj = sheet_obj.cell(row=v, column=16)
                cell_obj.font = Font(b=True, size=12, color="00FFFFFF")
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                cell_obj.fill = PatternFill("solid", fgColor="00003366")

            ## creacion de VLOOKUP para el nombre
            for vl in range(3, max_r + 2):
                vlookup4 = "=VLOOKUP(C" + str(vl) + ",PPG_DB!A:B,2,FALSE)"
                cell_obj = sheet_obj.cell(row=vl, column=18, value=vlookup4)
                cell_obj.font = Font(b=True, size=12)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #### VLOOKUP para el destino
            for vl in range(3, max_r + 2):
                vlookup5 = "=VLOOKUP(C" + str(vl) + ",PPG_DB!A:C,3,FALSE)"
                cell_obj = sheet_obj.cell(row=vl, column=19, value=vlookup5)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            ## creacion de VLOOKUP para el Fraccion Arancelaria [AF]
            for vl in range(3, max_r + 2):
                vlookup6 = "=VLOOKUP(D" + str(vl) + ",Catalogo!A:C,2,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=32, value=vlookup6)
                cell_obj.font = Font(size=12)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                vlookup6 = "=VLOOKUP(D" + str(vl) + ",Catalogo!A:C,3,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=33, value=vlookup6)
                cell_obj.font = Font(size=12)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                vlookup6 = "=VLOOKUP(D" + str(vl) + ",Catalogo!A:F,6,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=21, value=vlookup6)
                cell_obj.font = Font(size=12)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #### VLOOKUP para el el area de SAT [V y W]
            for vl in range(3, max_r + 2):
                vlookup3 = "=VLOOKUP(U" + str(vl) + ",c_ClaveProdServCP!A:C,2,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=22, value=vlookup3)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                vlookup3 = "=VLOOKUP(U" + str(vl) + ",c_ClaveProdServCP!A:C,3,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=23, value=vlookup3)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #### VLOOKUP para el el area de UN importaciones [Y]
            for vl in range(3, max_r + 2):
                vlookup7 = "=VLOOKUP(D" + str(vl) + ",UN_Importaciones!B:D,2,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=25, value=vlookup7)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                vlookup7 = "=VLOOKUP(D" + str(vl) + ",UN_Importaciones!B:E,4,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=26, value=vlookup7)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                vlookup7 = "=VLOOKUP(D" + str(vl) + ",UN_Importaciones!B:D,3,TRUE)"
                cell_obj = sheet_obj.cell(row=vl, column=27, value=vlookup7)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")


            ###### Segunda seccion ####
            sheet_obj.cell(row=max_r + 10, column=5, value="PIPAS Y TOTES A CLIENTES")
            for head2 in range(1, max_col + 6):
                cell_obj = sheet_obj.cell(row=max_r + 10, column=head2)
                cell_obj.font = Font(b=True, size=22, color="ffffff")
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                cell_obj.fill = PatternFill("solid", fgColor="960000")

            #------------------------------------------------------------------------------------#
            ###### crear una segunda hoja para datos pivote
            sheet_obj = wb_obj.create_sheet("PPG_DB")
            sheet_obj2 = wb_obj['PPG_DB']

            # incremento de columnas en db
            sheet_obj2.column_dimensions['A'].width = 45
            sheet_obj2.column_dimensions['B'].width = 45
            sheet_obj2.column_dimensions['C'].width = 45

            ### pegar info de DB a PPG_db
            for db in range(1, max_r_db + 2):
                c_value = sheet_obj_db.cell(row=db, column=2).value
                cell_obj = sheet_obj2.cell(row=db, column=1, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db in range(1, max_r_db + 2):
                c_value = sheet_obj_db.cell(row=db, column=12).value
                cell_obj = sheet_obj2.cell(row=db, column=2, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db in range(1, max_r_db + 2):
                c_value = sheet_obj_db.cell(row=db, column=18).value
                cell_obj = sheet_obj2.cell(row=db, column=3, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #------------------------------------------------------------------------------------#
            ###### crear una tercera hoja para datos de Catalogo pivote [c_ClaveProdServCP]
            sheet_obj = wb_obj.create_sheet("c_ClaveProdServCP")
            sheet_obj3 = wb_obj['c_ClaveProdServCP']

            # incremento de columnas en db
            sheet_obj3.column_dimensions['A'].width = 45
            sheet_obj3.column_dimensions['B'].width = 45
            sheet_obj3.column_dimensions['C'].width = 45

            ### pegar info del archivo [c_ClaveProdServCP.xlsx] a hoja c_ClaveProdServCP
            for db3 in range(6, max_r_db3 + 2):
                c_value = sheet_obj_db3.cell(row=db3, column=1).value
                cell_obj = sheet_obj3.cell(row=db3, column=1, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db3 in range(6, max_r_db3 + 2):
                c_value = sheet_obj_db3.cell(row=db3, column=2).value
                cell_obj = sheet_obj3.cell(row=db3, column=2, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db3 in range(6, max_r_db3 + 2):
                c_value = sheet_obj_db3.cell(row=db3, column=4).value
                cell_obj = sheet_obj3.cell(row=db3, column=3, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            #------------------------------------------------------------------------------------#
            ###### crear una cuarta hoja para datos de Catalogo pivote [c_MaterialPeligroso]
            sheet_obj = wb_obj.create_sheet("c_MaterialPeligroso")
            sheet_obj4 = wb_obj['c_MaterialPeligroso']

            # incremento de columnas en db
            sheet_obj4.column_dimensions['A'].width = 45
            sheet_obj4.column_dimensions['B'].width = 45
            sheet_obj4.column_dimensions['C'].width = 45

            #------------------------------------------------------------------------------------#
            ###### crear una quinta hoja para datos de Catalogo pivote [UN_Importaciones]
            sheet_obj = wb_obj.create_sheet("UN_Importaciones")
            sheet_obj5 = wb_obj['UN_Importaciones']

            # incremento de columnas en db
            sheet_obj5.column_dimensions['A'].width = 45
            sheet_obj5.column_dimensions['B'].width = 45
            sheet_obj5.column_dimensions['C'].width = 45
            sheet_obj5.column_dimensions['D'].width = 45
            sheet_obj5.column_dimensions['E'].width = 45
            sheet_obj5.column_dimensions['F'].width = 45

            ### pegar info del archivo [un_importaciones.xlsx] a hoja un_importaciones
            for db5 in range(1, max_r_db5 + 2):
                c_value = sheet_obj_db5.cell(row=db5, column=2).value
                cell_obj = sheet_obj5.cell(row=db5, column=2, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                c_value = sheet_obj_db5.cell(row=db5, column=3).value
                cell_obj = sheet_obj5.cell(row=db5, column=3, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                c_value = sheet_obj_db5.cell(row=db5, column=4).value
                cell_obj = sheet_obj5.cell(row=db5, column=4, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")
                c_value = sheet_obj_db5.cell(row=db5, column=5).value
                cell_obj = sheet_obj5.cell(row=db5, column=5, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")



            # ------------------------------------------------------------------------------------#
            ###### crear una quinta hoja para datos de Catalogo pivote [Catalogo]
            sheet_obj = wb_obj.create_sheet("Catalogo")
            sheet_obj6 = wb_obj['Catalogo']

            # incremento de columnas en db
            sheet_obj6.column_dimensions['A'].width = 45
            sheet_obj6.column_dimensions['B'].width = 45
            sheet_obj6.column_dimensions['C'].width = 45
            sheet_obj6.column_dimensions['D'].width = 45

            ### pegar info del archivo [c_ClaveProdServCP.xlsx] a hoja c_ClaveProdServCP
            for db6 in range(1, max_r_db6 + 2):
                c_value = sheet_obj_db6.cell(row=db6, column=1).value
                cell_obj = sheet_obj6.cell(row=db6, column=1, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db6 in range(1, max_r_db6 + 2):
                c_value = sheet_obj_db6.cell(row=db6, column=2).value
                cell_obj = sheet_obj6.cell(row=db6, column=2, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db6 in range(1, max_r_db6 + 2):
                c_value = sheet_obj_db6.cell(row=db6, column=3).value
                cell_obj = sheet_obj6.cell(row=db6, column=3, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")

            for db6 in range(1, max_r_db6 + 2):
                c_value = sheet_obj_db6.cell(row=db6, column=6).value
                cell_obj = sheet_obj6.cell(row=db6, column=6, value=c_value)
                cell_obj.alignment = Alignment(horizontal="center", vertical="center")



            # guarda el archivo
            wb_obj.save(export)
            exit()
